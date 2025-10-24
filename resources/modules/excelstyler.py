"""Módulo para estilização de planilhas Excel.

Este módulo fornece utilitários para aplicar estilos e formatação em
planilhas Excel usando openpyxl.
"""

import logging
from typing import List, Optional

from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


class ExcelFormatter:
    """Formatador de planilhas Excel.

    Fornece métodos estáticos para aplicar estilos, alinhamentos e bordas
    em planilhas Excel.

    Example:
        >>> from openpyxl import load_workbook
        >>> wb = load_workbook('data.xlsx')
        >>> ws = wb.active
        >>> ExcelFormatter.adjust_column_width(ws)
        >>> ExcelFormatter.center_text_in_cells(ws)
        >>> wb.save('data_formatted.xlsx')
    """

    @staticmethod
    def adjust_column_width(worksheet: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
        """Ajusta a largura das colunas baseado no conteúdo das células.

        Args:
            worksheet: Planilha do Excel a ser ajustada.
            min_width: Largura mínima da coluna (padrão: 10).
            max_width: Largura máxima da coluna (padrão: 50).

        Example:
            >>> ExcelFormatter.adjust_column_width(worksheet)
        """
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = None

                for cell in column:
                    column_letter = cell.column_letter
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                    except (AttributeError, TypeError) as e:
                        logging.debug(f"Erro ao calcular comprimento da célula: {e}")
                        continue

                if column_letter:
                    # Ajusta largura com padding e limites
                    adjusted_width = min(max(max_length + 2, min_width), max_width)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            logging.error(f"Erro ao ajustar largura das colunas: {e}")
            raise

    @staticmethod
    def adjust_alignment_in_column(
        worksheet: Worksheet,
        column_letter: str,
        horizontal: str = 'center',
        vertical: str = 'center',
        wrap_text: bool = True
    ) -> None:
        """Ajusta o alinhamento das células em uma coluna específica.

        Args:
            worksheet: Planilha do Excel a ser ajustada.
            column_letter: Letra da coluna a ser ajustada (ex: 'A', 'B').
            horizontal: Alinhamento horizontal ('left', 'center', 'right').
            vertical: Alinhamento vertical ('top', 'center', 'bottom').
            wrap_text: Se True, habilita quebra de texto.

        Example:
            >>> ExcelFormatter.adjust_alignment_in_column(worksheet, 'A', 'left')
        """
        try:
            alignment = Alignment(
                horizontal=horizontal,
                vertical=vertical,
                wrap_text=wrap_text
            )

            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.column_letter == column_letter:
                        cell.alignment = alignment

        except Exception as e:
            logging.error(f"Erro ao ajustar alinhamento da coluna {column_letter}: {e}")
            raise

    @staticmethod
    def center_text_in_cells(
        worksheet: Worksheet,
        exclude_columns: Optional[List[str]] = None
    ) -> None:
        """Centraliza o texto em todas as células da planilha.

        Args:
            worksheet: Planilha do Excel a ser ajustada.
            exclude_columns: Lista de letras de colunas a serem excluídas (ex: ['A', 'C']).

        Example:
            >>> ExcelFormatter.center_text_in_cells(worksheet)
            >>> ExcelFormatter.center_text_in_cells(worksheet, exclude_columns=['A'])
        """
        try:
            exclude_columns = exclude_columns or []
            alignment = Alignment(horizontal='center', vertical='center')

            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.column_letter not in exclude_columns:
                        cell.alignment = alignment

        except Exception as e:
            logging.error(f"Erro ao centralizar texto nas células: {e}")
            raise

    @staticmethod
    def add_borders_to_cells(
        worksheet: Worksheet,
        border_style: str = 'thin'
    ) -> None:
        """Adiciona bordas a todas as células da planilha.

        Args:
            worksheet: Planilha do Excel a ser ajustada.
            border_style: Estilo da borda ('thin', 'medium', 'thick').

        Example:
            >>> ExcelFormatter.add_borders_to_cells(worksheet)
            >>> ExcelFormatter.add_borders_to_cells(worksheet, 'medium')
        """
        try:
            border = Border(
                left=Side(style=border_style),
                right=Side(style=border_style),
                top=Side(style=border_style),
                bottom=Side(style=border_style)
            )

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = border

        except Exception as e:
            logging.error(f"Erro ao adicionar bordas às células: {e}")
            raise

    @staticmethod
    def apply_header_style(
        worksheet: Worksheet,
        row_number: int = 1,
        bold: bool = True,
        bg_color: Optional[str] = None
    ) -> None:
        """Aplica estilo especial à linha de cabeçalho.

        Args:
            worksheet: Planilha do Excel a ser ajustada.
            row_number: Número da linha do cabeçalho (padrão: 1).
            bold: Se True, aplica negrito.
            bg_color: Cor de fundo em hexadecimal (ex: 'CCCCCC').

        Example:
            >>> ExcelFormatter.apply_header_style(worksheet, bg_color='CCCCCC')
        """
        try:
            from openpyxl.styles import Font, PatternFill

            for cell in worksheet[row_number]:
                if bold:
                    cell.font = Font(bold=True)

                if bg_color:
                    cell.fill = PatternFill(
                        start_color=bg_color,
                        end_color=bg_color,
                        fill_type='solid'
                    )

                cell.alignment = Alignment(horizontal='center', vertical='center')

        except Exception as e:
            logging.error(f"Erro ao aplicar estilo ao cabeçalho: {e}")
            raise


# Mantém compatibilidade com código legado
class ExcelStyler(ExcelFormatter):
    """Classe de compatibilidade com código legado.

    Deprecated: Use ExcelFormatter diretamente.
    """

    @staticmethod
    def center_text_in_cells_except(worksheet: Worksheet, exclude_columns: List[str] = None) -> None:
        """Método legado - use center_text_in_cells().

        Deprecated: Use ExcelFormatter.center_text_in_cells(worksheet, exclude_columns) diretamente.
        """
        exclude_columns = exclude_columns or []
        return ExcelFormatter.center_text_in_cells(worksheet, exclude_columns)
