import logging
import os
import shutil
import ssl
import time
import urllib.request
from datetime import datetime
from typing import Dict, List, Optional, Union

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


class File:
    """Classe responsável por operações relacionadas a arquivos."""

    @staticmethod
    def _create_directory_if_not_exists(directory: str) -> None:
        """
        Cria um diretório se ele não existir.

        Args:
            directory (str): Caminho do diretório a ser criado.
        """
        os.makedirs(directory, exist_ok=True)

    @staticmethod
    def download_url(url: str, directory: str) -> None:
        """
        Baixa um arquivo da URL fornecida e o salva no diretório especificado.

        Args:
            url (str): URL do arquivo a ser baixado.
            directory (str): Caminho do diretório onde o arquivo será salvo.

        Raises:
            FileNotFoundError: Se ocorrer um erro ao baixar o arquivo.
        """
        try:
            logging.debug(f"Downloading file from {url} to {directory}")
            
            File._create_directory_if_not_exists(directory=directory)

            file_path = os.path.join(directory, os.path.basename(url))

            if os.path.isfile(file_path):
                os.remove(file_path)

            ssl_context = ssl.create_default_context()
            ssl_context.check_hostname = False
            ssl_context.verify_mode = ssl.CERT_NONE

            with urllib.request.urlopen(url, context=ssl_context) as response, open(file_path, 'wb') as out_file:
                out_file.write(response.read())

        except Exception as error:
            raise FileNotFoundError(f"Erro ao baixar o arquivo {url}: {error}")

    @staticmethod
    def wait_for_download(path_origem: str, timeout: int = 15, check_interval: int = 3) -> None:
        """
        Aguarda até que um arquivo no diretório especificado termine de ser baixado.

        A função verifica se o tamanho do arquivo permanece inalterado durante o intervalo de verificação
        para determinar se o download foi concluído.

        Args:
            path_origem (str): Caminho do diretório onde o arquivo está sendo baixado.
            timeout (int, optional): Tempo máximo em segundos para aguardar pelo término do download. Padrão é 15 segundos.
            check_interval (int, optional): Intervalo em segundos para verificar o tamanho do arquivo. Padrão é 3 segundos.

        Returns:
            None: Retorna None se o download for concluído com sucesso dentro do timeout.
            Se o tempo limite for excedido, um aviso será registrado.
        """
        inicio = time.time()
        previous_size = -1
        
        while time.time() - inicio < timeout:
            if os.listdir(path_origem):
                file_path = os.path.join(path_origem, os.listdir(path_origem)[0])
                current_size = os.path.getsize(file_path)
                
                if current_size == previous_size:
                    return
                
                previous_size = current_size
            time.sleep(check_interval)
        else:
            logging.warning("Tempo limite excedido ao esperar pelo download")

    @staticmethod
    def wait_for_files(path: str, ext: str, timeout: int = 10) -> Union[List[str], bool]:
        """
        Aguarda por arquivos com uma extensão específica em um diretório por um tempo limite.

        Args:
            path (str): Caminho do diretório onde os arquivos serão procurados.
            ext (str): Extensão dos arquivos que devem ser esperados.
            timeout (int, optional): Tempo limite para aguardar os arquivos em segundos. Padrão: 10 segundos.

        Returns:
            Union[List[str], bool]: Lista de arquivos encontrados com a extensão desejada ou False se não encontrado.
        """
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            try:
                files = os.listdir(path)
                filtered_files = [file for file in files if file.upper().endswith(ext.upper())]
                
                if filtered_files:
                    return filtered_files
                
                time.sleep(1)
            except Exception as e:
                logging.warning("Error waiting for file download: " + str(e))
                return False
        
        return False

    @staticmethod
    def move_files_by_extension(source_path: str, destination_path: str, extension: str) -> None:
        """
        Move arquivos com uma determinada extensão de um diretório de origem para um destino.

        Args:
            source_path (str): Caminho do diretório de origem.
            destination_path (str): Caminho do diretório de destino.
            extension (str): Extensão dos arquivos a serem movidos (ex: ".pdf").

        Raises:
            Exception: Se ocorrer um erro ao mover os arquivos.
        """
        try:
            logging.info("Movendo arquivos!")
            File._create_directory_if_not_exists(destination_path)

            for file in os.listdir(source_path):
                if file.upper().endswith(extension.upper()):
                    shutil.move(os.path.join(source_path, file), destination_path)
        except Exception as e:
            logging.warning(f"Erro ao mover arquivos: {e}")
            raise Exception(f"Erro ao mover arquivos: {e}") from e

    @staticmethod
    async def clear_folder(source_path: str) -> None:
        """
        Deleta todos os arquivos e diretórios dentro do diretório especificado.

        Args:
            source_path (str): Caminho do diretório a ser limpo.

        Raises:
            Exception: Se ocorrer um erro ao deletar os arquivos.
        """
        try:
            logging.info("Deletando arquivos!")
            File._create_directory_if_not_exists(source_path)

            with os.scandir(source_path) as entries:
                for entry in entries:
                    if entry.is_file():
                        os.remove(entry.path)
                    elif entry.is_dir():
                        os.rmdir(entry.path)
        except Exception as e:
            logging.warning(f"Erro ao deletar arquivos: {e}")
            raise Exception(f"Erro ao deletar arquivos: {e}") from e

    @staticmethod
    def delete_files_extension(directory: str, extension: str) -> None:
        """
        Deleta todos os arquivos com uma determinada extensão em um diretório especificado.

        Args:
            directory (str): Caminho do diretório onde os arquivos serão deletados.
            extension (str): Extensão dos arquivos a serem deletados (ex: ".pdf").
        """
        for filename in os.listdir(directory):
            if filename.endswith(extension):
                file_path = os.path.join(directory, filename)
                os.remove(file_path)

    @staticmethod
    def delete_file(filename: str) -> None:
        """
        Deleta um arquivo especificado.

        Args:
            filename (str): Caminho do arquivo a ser deletado.

        Raises:
            OSError: Se ocorrer um erro ao deletar o arquivo.
        """
        try:
            os.remove(filename)
        except OSError as error:
            raise OSError(f"Erro ao deletar o arquivo {filename}: {error}")
        
    @staticmethod
    def delete_files_directory(directory: str) -> None:
        """
        Deleta todos os arquivos em um diretório especificado, mas mantém o diretório.

        Args:
            directory (str): Caminho do diretório onde os arquivos serão deletados.
        """
        logging.debug(f"Cleaning the directory {directory}")
        File._create_directory_if_not_exists(directory)
        for filename in os.listdir(directory):
            file_path = os.path.join(directory, filename)
            os.remove(file_path)

    @staticmethod
    def delete_files_and_directory(directory: str) -> None:
        """
        Deleta todos os arquivos em um diretório especificado e em seguida deleta o próprio diretório.

        Args:
            directory (str): Caminho do diretório onde os arquivos serão deletados.
        """
        if os.path.exists(directory):
            for filename in os.listdir(directory):
                file_path = os.path.join(directory, filename)
                os.remove(file_path)
            os.rmdir(directory)

    @staticmethod
    def check_files_exist(destination_path: str, prefix: str) -> bool:
        """
        Verifica se existem arquivos com um determinado prefixo no diretório especificado.

        Args:
            destination_path (str): Caminho do diretório onde a verificação será realizada.
            prefix (str): Prefixo dos arquivos a serem verificados.

        Returns:
            bool: True se existir pelo menos um arquivo com o prefixo especificado, False caso contrário.
        """
        files = [f for f in os.listdir(destination_path) if os.path.isfile(os.path.join(destination_path, f))]
        return any(file.lower().startswith(prefix.lower()) for file in files)
    
    @staticmethod
    def get_files_directory(directory: str) -> List[str]:
        """
        Retorna todos os arquivos de um diretório especificado.

        Args:
            directory (str): Caminho do diretório do qual os arquivos serão listados.

        Returns:
            List[str]: Lista contendo o nome de todos os arquivos no diretório especificado.
        """
        logging.debug(f"Listing files in directory {directory}")
        return [arquivo for arquivo in os.listdir(directory) if os.path.isfile(os.path.join(directory, arquivo))]
    
    @staticmethod
    def create_xlsx_with_tabs(tab_names: dict, file_name: str, directory: str):
        """
        Cria um arquivo xlsx com as abas e colunas especificadas.
        
        Args:
            tab_names (dict): Dicionário contendo os nomes das abas como chaves e as colunas como valores.
            file_name (str): Nome do arquivo a ser criado.
            directory (str): Caminho do diretório onde o arquivo será criado.
        """
        logging.debug(f"Creating xlsx file with tabs {list(tab_names.keys())}")
        wb = Workbook()
        
        for index, (tab_name, columns) in enumerate(tab_names.items()):
            if index == 0:
                ws = wb.active
                ws.title = tab_name
            else:
                ws = wb.create_sheet(title=tab_name)
            
            for col_num, col_name in enumerate(columns, 1):
                ws.cell(row=1, column=col_num, value=col_name)

        wb.save(os.path.join(directory, file_name))
    
    @staticmethod
    def read_excel(file) -> pd.DataFrame:
        """
        Lê o arquivo xls ou xlsx do diretório fornecido e retorna um DataFrame.

        Args:
            file (str): Caminho completo do arquivo

        Returns:
            pandas.DataFrame: DataFrame com os dados do arquivo xls ou xlsx
        """
        file_extension = (os.path.splitext(file)[1]).lower()
        
        if file_extension == '.xls':
            df = pd.read_excel(file, engine='xlrd')
        elif file_extension == '.xlsx':
            df = pd.read_excel(file, engine='openpyxl')
        else:
            raise Exception(f'Extensão do arquivo {file} não é suportada')
        
        return df
    
    @staticmethod
    def move_file_to_backup(file: str, dir_backup: str, name_file: str = "Arquivo_Original") -> None:
        """
        Move o arquivo para a pasta de backup.
        
        Args:
            file (str): Caminho completo do arquivo.
            dir_backup (str): Caminho da pasta de backup.
            name_file (str, optional): Nome prefixo para o arquivo a ser movido. Padrão: "Arquivo_Original".
        """
        logging.debug(f"Moving file {file} to backup folder")
        File._create_directory_if_not_exists(dir_backup)
        file_extension = os.path.splitext(file)[1]
        base_name_without_extension = os.path.splitext(os.path.basename(file))[0]
        date_time = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        file_name_destin = f"{name_file}_{base_name_without_extension}_{date_time}{file_extension}"
        shutil.move(file, os.path.join(dir_backup, file_name_destin)) 